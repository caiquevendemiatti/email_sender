from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def gerar_workbook_contatos(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contatos'

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    id_header_fill = PatternFill(start_color='595959', end_color='595959', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    cell_font = Font(name='Calibri', size=10)
    id_cell_font = Font(name='Calibri', size=10, color='595959')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['ID', 'Razão Social', 'Contato', 'E-mail', 'Ativo', 'Excluído', 'Colaborador Responsável']
    col_widths = [8, 45, 35, 45, 10, 12, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = id_header_fill if col_idx == 1 else header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20
    alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    for row_idx, contato in enumerate(queryset, start=2):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            contato.pk,
            contato.razao_social,
            contato.contato,
            contato.e_mail,
            'Sim' if contato.ativo else 'Não',
            'Sim' if contato.excluido else 'Não',
            contato.colaborador_responsavel.nome,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = id_cell_font if col_idx == 1 else cell_font
            cell.border = border
            cell.alignment = center if col_idx in (1, 5, 6) else left
            if row_fill:
                cell.fill = row_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    return wb
