import re
from openpyxl import load_workbook
from django.db import transaction
from emails_controller.models import Contato, Colaborador


SIM_VALUES = {'sim', 's', '1', 'true', 'yes'}

REQUIRED_COLUMNS = {'ID', 'Razão Social', 'Contato', 'E-mail', 'Ativo', 'Excluído', 'Colaborador Responsável'}


def parse_bool(value):
    return str(value).strip().lower() in SIM_VALUES


def clean_email(raw):
    email = str(raw).strip().strip(';,').strip()
    if re.search(r'[;,]', email):
        return None, f'E-mail contém múltiplos endereços ou caractere inválido: "{raw}"'
    return email, None


def processar_importacao(file_obj, dry_run=False):
    """
    Processa um arquivo Excel de contatos (upload ou caminho).
    Retorna dict com: total, created, updated, skipped, errors (lista de dicts row/message).
    """
    result = {'total': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        result['errors'].append({'row': None, 'message': 'Planilha vazia.'})
        return result

    header = [str(h).strip() if h else '' for h in rows[0]]
    missing = REQUIRED_COLUMNS - set(header)
    if missing:
        result['errors'].append({'row': None, 'message': f'Colunas ausentes na planilha: {missing}'})
        return result

    col = {name: idx for idx, name in enumerate(header)}
    colaboradores_cache = {c.nome: c for c in Colaborador.objects.all()}

    for row_num, row in enumerate(rows[1:], start=2):
        result['total'] += 1

        contact_id = row[col['ID']]
        razao_social = str(row[col['Razão Social']] or '').strip()
        contato_nome = str(row[col['Contato']] or '').strip()
        email, email_erro = clean_email(row[col['E-mail']] or '')
        ativo = parse_bool(row[col['Ativo']])
        excluido = parse_bool(row[col['Excluído']])
        colaborador_nome = str(row[col['Colaborador Responsável']] or '').strip()

        if email_erro:
            result['errors'].append({'row': row_num, 'message': f'{email_erro} — corrija na planilha e reimporte.'})
            continue

        colaborador = colaboradores_cache.get(colaborador_nome)
        if not colaborador:
            result['errors'].append({'row': row_num, 'message': f'Colaborador "{colaborador_nome}" não encontrado.'})
            continue

        # Novo contato (ID vazio)
        if not contact_id:
            if not razao_social or not email:
                result['errors'].append({'row': row_num, 'message': 'Novo contato sem Razão Social ou E-mail.'})
                continue

            if not dry_run:
                try:
                    with transaction.atomic():
                        Contato.objects.create(
                            razao_social=razao_social,
                            contato=contato_nome,
                            e_mail=email,
                            ativo=ativo,
                            excluido=excluido,
                            colaborador_responsavel=colaborador,
                        )
                except Exception as e:
                    result['errors'].append({'row': row_num, 'message': f'Erro ao criar "{razao_social}": {e}'})
                    continue

            result['created'] += 1
            continue

        # Atualização (ID preenchido)
        try:
            contato = Contato.objects.select_related('colaborador_responsavel').get(pk=int(contact_id))
        except Contato.DoesNotExist:
            result['errors'].append({'row': row_num, 'message': f'Contato ID={contact_id} não encontrado — verifique o ID.'})
            continue

        changed = (
            contato.razao_social != razao_social
            or contato.contato != contato_nome
            or contato.e_mail != email
            or contato.ativo != ativo
            or contato.excluido != excluido
            or contato.colaborador_responsavel_id != colaborador.pk
        )

        if not changed:
            result['skipped'] += 1
            continue

        if not dry_run:
            try:
                with transaction.atomic():
                    contato.razao_social = razao_social
                    contato.contato = contato_nome
                    contato.e_mail = email
                    contato.ativo = ativo
                    contato.excluido = excluido
                    contato.colaborador_responsavel = colaborador
                    contato.save()
            except Exception as e:
                result['errors'].append({'row': row_num, 'message': f'Erro ao atualizar ID={contact_id}: {e}'})
                continue

        result['updated'] += 1

    return result


def formatar_resultado(result, dry_run=False):
    prefix = '[SIMULAÇÃO] ' if dry_run else ''
    lines = [
        f'{prefix}Total processado: {result["total"]}',
        f'Criados: {result["created"]}',
        f'Atualizados: {result["updated"]}',
        f'Sem alterações: {result["skipped"]}',
        f'Erros: {len(result["errors"])}',
    ]
    if result['errors']:
        lines.append('\nDetalhes dos erros:')
        for e in result['errors']:
            row_info = f'Linha {e["row"]}: ' if e['row'] else ''
            lines.append(f'  • {row_info}{e["message"]}')
    return '\n'.join(lines)
